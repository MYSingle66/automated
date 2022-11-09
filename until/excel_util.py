# -*- coding:utf-8 -*-
# @Time: 2022/9/27 16:24
# @Author: 小杨同学
# @Email: 541667133@qq.com
# @File: excel_util.py
import json
import openpyxl
def read_data_all(filename, sheetname):
    """
    :param filename: 文件名
    :param sheetname: sheet名
    :return:
    """
    book = openpyxl.load_workbook(filename)
    sh1 = book[sheetname]
    all_case_data=[]
    rows = sh1.max_row
    print(rows)
    for i in range(2, rows + 1, 1):
        caseid = sh1.cell(i, 1).value+sh1.cell(i, 2).value
        url = sh1.cell(i, 4).value
        method = sh1.cell(i, 5).value
        headers = sh1.cell(i, 6).value
        data = sh1.cell(i, 7).value
        excepted_res = sh1.cell(i, 8).value
        cate_data = dict(dcaseid=caseid,dmethod=method,dheaders={"Content-Type": headers},durl=url,ddate=eval(data),dexcepted_res=eval(excepted_res))
        all_case_data.append(cate_data)
    # reset=json.dumps(all_case_data, sort_keys=True, indent=4, separators=(',', ': ') ,ensure_ascii=False)
    # print(reset)
    return all_case_data

def write_res(filename,sheetname,row,fina_res):
    """
    :param filename: 文件名
    :param sheetname: sheet名
    :param row: 行数
    :param fina_res: 实际结果
    :return:
    """
    book = openpyxl.load_workbook(filename)
    sh1 = book[sheetname]
    sh1.cell(row,9).value=fina_res
    book.save(filename)

if __name__ == '__main__':
    read_data_all("E:/python/Exceldemo/apitest.xlsx", "Sheet1")
